# Standard library imports
import asyncio
import csv
import os
import re
import time
import zipfile
from enum import Enum
from typing import Any, Dict, Union
from uuid import UUID

# Third party imports
import aiofiles
import aiofiles.os as aos
import pandas as pd
from fastapi import HTTPException, status
from openpyxl import load_workbook
from osgeo import ogr, osr
from pydantic import BaseModel, HttpUrl
from pyproj import CRS
from qgis.core import (
    QgsProject,
    QgsVectorFileWriter,
    QgsVectorLayer,
)
from shapely import wkb
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.sql import select, text
from sqlmodel import SQLModel
from starlette.datastructures import UploadFile

# Local application imports
from core.core.config import settings
from core.core.job import job_log
from core.crud.base import CRUDBase
from core.db.models._link_model import LayerProjectLink
from core.db.models.layer import (
    FeatureDataType,
    FeatureLayerExportType,
    FeatureType,
    FileUploadType,
    Layer,
    LayerType,
    TableLayerExportType,
)
from core.schemas.error import DataOutCRSBoundsError, Ogr2OgrError
from core.schemas.job import JobStatusType, Msg, MsgType
from core.schemas.layer import (
    IFileUploadExternalService,
    MaxFileSizeType,
    NumberColumnsPerType,
    OgrDriverType,
    OgrPostgresSubType,
    OgrPostgresType,
    SupportedOgrGeomType,
)
from core.utils import (
    async_delete_dir,
    async_run_command,
    async_scandir,
    print_warning,
    sanitize_error_message,
)


async def delete_old_files(max_time: int) -> None:
    """Delete old files from data directory."""
    # Clean all old folders that are older then two hours
    async for folder in async_scandir(settings.DATA_DIR):
        stat_result = await aos.stat(os.path.join(settings.DATA_DIR, folder.name))
        if stat_result.st_mtime < (time.time() - 2 * 3600):
            await async_delete_dir(os.path.join(settings.DATA_DIR, folder.name))


def model_to_dict(model: SQLModel | BaseModel) -> Dict[str, Any]:
    model_dict = model.model_dump()
    for key, value in model_dict.items():
        if isinstance(value, Enum):
            model_dict[key] = value.value
    return model_dict


def get_user_table(layer: Union[dict, SQLModel, BaseModel]) -> str:
    """Get the table with the user data based on the layer metadata."""

    # Check if layer is of type dict or SQLModel/BaseModel
    if isinstance(layer, (SQLModel, BaseModel)):
        layer = model_to_dict(layer)

    if isinstance(layer, dict):
        if layer["type"] == LayerType.feature.value:
            if layer["feature_layer_type"] in (FeatureType.standard, FeatureType.tool):
                table_prefix = layer["feature_layer_geometry_type"]
            elif layer["feature_layer_type"] == FeatureType.street_network:
                table_prefix = (
                    FeatureType.street_network.value
                    + "_"
                    + layer["feature_layer_geometry_type"]
                )
        elif layer["type"] == LayerType.table.value:
            table_prefix = "no_geometry"
        else:
            raise ValueError(f"The passed layer type {layer['type']} is not supported.")
    user_id = layer["user_id"]
    return f"{settings.USER_DATA_SCHEMA}.{table_prefix}_{str(user_id).replace('-', '')}"


class CRUDLayerBase(CRUDBase):
    async def check_and_alter_layer_name(
        self,
        async_session: AsyncSession,
        folder_id: UUID,
        layer_name: str,
        project_id: UUID | None = None,
    ) -> str:
        """Check if layer name already exists in project and alter it like layer (n+1) if necessary"""

        # Regular expression to find layer names with a number
        pattern = re.compile(rf"^{re.escape(layer_name)} \((\d+)\)$")

        # Get all layer names in project
        if project_id:
            names_in_project = await async_session.execute(
                select(LayerProjectLink.name).where(
                    LayerProjectLink.project_id == project_id,
                    LayerProjectLink.name.like(f"{layer_name}%"),
                )
            )
            layer_names = [row[0] for row in names_in_project.fetchall()]
        else:
            layer_names = []

        # Get all layer names in folder
        names_in_folder = [
            row[0]
            for row in (
                await async_session.execute(
                    select(Layer.name).where(
                        Layer.folder_id == folder_id,
                        Layer.name.like(f"{layer_name}%"),
                    )
                )
            ).fetchall()
        ]
        layer_names = list(set(layer_names + names_in_folder))

        # Find the highest number (n) among the layer names using list comprehension
        numbers = [
            int(match.group(1))
            for name in layer_names
            if (match := pattern.match(name))
        ]
        highest_num = max(numbers, default=0)

        # Check if the base layer name exists
        base_name_exists = layer_name in layer_names

        # Construct the new layer name
        if base_name_exists or highest_num > 0:
            new_layer_name = f"{layer_name} ({highest_num + 1})"
        else:
            new_layer_name = layer_name

        return new_layer_name


class FileUpload:
    def __init__(
        self,
        async_session: AsyncSession,
        user_id: UUID,
        dataset_id: UUID,
        source: UploadFile | IFileUploadExternalService,
    ) -> None:
        self.async_session = async_session
        self.user_id = user_id
        self.source = source
        self.folder_path = os.path.join(
            settings.DATA_DIR, str(self.user_id), str(dataset_id)
        )

        if isinstance(source, UploadFile):
            self.file_ending = os.path.splitext(source.filename)[-1][1:]
            self.file_path = os.path.join(self.folder_path, "file." + self.file_ending)
        else:
            self.file_path = os.path.join(
                self.folder_path, "file." + FileUploadType.geojson.value
            )

    async def _fetch_and_write(self) -> None:
        """Fetch data from external service if required, save file to disk."""

        if isinstance(self.source, UploadFile):
            # An existing file was uploaded, save file in chunks
            async with aiofiles.open(self.file_path, "wb") as buffer:
                while True:
                    chunk = await self.source.read(65536)
                    if not chunk:
                        break
                    await buffer.write(chunk)
        else:
            # Ensure a URL is specified
            url = self.source.other_properties.url
            if not url:
                raise ValueError(
                    "A URL must be specified under the 'url' key of other_properties."
                )

            # Initialize OGR external service fetching
            fetch_layer_external_service = FetchLayerExternalService(
                url=url, output_file=self.file_path
            )

            if self.source.data_type == FeatureDataType.wfs:
                # Ensure a single WFS layer is specified
                layers = self.source.other_properties.layers
                if not layers or len(layers) != 1:
                    raise ValueError(
                        "WFS: A single layer must be specified under the 'layers' key of other_properties."
                    )

                # Fetch the specified layer
                fetch_layer_external_service.fetch_wfs(layer_name=layers[0])

                # Ensure the newly saved file does not exceed file size limits
                if (
                    os.path.getsize(self.file_path)
                    > MaxFileSizeType[FileUploadType.gpkg.value].value
                ):
                    raise HTTPException(
                        status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                        detail=f"File size too large. Max file size is {round(MaxFileSizeType[FileUploadType.gpkg.value].value / 1048576, 2)} MB",
                    )
            elif self.data_type == FeatureDataType.mvt:
                # TODO: Implement MVT fetching
                pass

    async def save_file(self) -> str:
        """Save file to disk for later operations."""

        # Clean all old folders that are older then two hours
        await delete_old_files(max_time=3600)

        # Create folder if exist delete it
        await async_delete_dir(self.folder_path)
        if not os.path.exists(os.path.join(settings.DATA_DIR, str(self.user_id))):
            await aos.mkdir(os.path.join(settings.DATA_DIR, str(self.user_id)))
        await aos.mkdir(self.folder_path)

        # Fetch and save file
        await self._fetch_and_write()

        return self.file_path

    async def save_file_fail(self) -> None:
        """Delete folder if file upload fails."""
        await async_delete_dir(self.folder_path)


class FetchLayerExternalService:
    def __init__(self, url: HttpUrl, output_file: str) -> None:
        self.MAX_FEATURE_COUNT = 200000

        self.url = url
        self.output_file = output_file

        # Output driver to be used
        self.output_driver_type = OgrDriverType.geojson.value

    def fetch_wfs(self, layer_name: str) -> None:
        """Fetch data from WFS service and save to disk."""

        # First, attempt to fetch data using QGIS
        try:
            # Create vector layer containing features from the WFS source
            uri = f"typename='{layer_name}' url='{self.url}'"
            layer = QgsVectorLayer(uri, layer_name, "WFS")

            if not layer.isValid():
                raise ValueError(f"Unable to open layer: {layer_name}")

            # Ensure layer is not too large
            feature_count = layer.featureCount()
            if feature_count > self.MAX_FEATURE_COUNT:
                raise ValueError(
                    f"Layer {layer_name} contains too many features ({feature_count})."
                )

            # Add layer to project and write to GeoJSON file
            QgsProject.instance().addMapLayer(layer)
            options = QgsVectorFileWriter.SaveVectorOptions()
            options.driverName = self.output_driver_type
            error = QgsVectorFileWriter.writeAsVectorFormatV2(
                layer,
                self.output_file,
                QgsProject.instance().transformContext(),
                options,
            )
            # Remove layer from project
            QgsProject.instance().removeMapLayer(layer)

            if error[0] != QgsVectorFileWriter.NoError:
                raise Exception(f"Unable to write GeoJSON file: {error[1]}")

            return
        except Exception as e:
            print_warning("QGIS failed to fetch WFS data, falling back to OGR.")
            print_warning(f"QGIS error: {e}")

        # Second, attempt to fetch data using OGR
        ogr.UseExceptions()

        # Initialize output GeoJSON driver
        output_driver = ogr.GetDriverByName(self.output_driver_type)
        if output_driver is None:
            raise Exception(f"{self.output_driver_type} driver is not available.")

        # Create output data source
        self.output_data_source = output_driver.CreateDataSource(self.output_file)
        if self.output_data_source is None:
            raise Exception(
                f"Could not create output data source at {self.output_file}"
            )

        # Initialize WFS data source
        wfs_data_source = ogr.Open(f"WFS:{str(self.url)}")
        if wfs_data_source is None:
            raise ValueError(f"Could not open WFS service at {self.url}")

        # Get the specified layer
        input_layer = wfs_data_source.GetLayerByName(layer_name)
        if input_layer is None:
            raise ValueError(f"Could not find layer {layer_name} in WFS service.")

        # Ensure layer is not too large
        feature_count = input_layer.GetFeatureCount()
        if feature_count > self.MAX_FEATURE_COUNT:
            raise ValueError(
                f"Layer {layer_name} contains too many features ({feature_count})."
            )

        # Get the layer definition
        input_layer_defn = input_layer.GetLayerDefn()
        if input_layer_defn is None:
            raise ValueError(f"Could not get layer definition for {layer_name}.")

        # Process geometry field
        if input_layer_defn.GetGeomFieldCount() != 1:
            raise ValueError(
                f"Layer {layer_name} must contain exactly one geometry field."
            )

        geom_type = input_layer_defn.GetGeomFieldDefn(0).GetType()
        if geom_type == ogr.wkbUnknown:
            first_feature = input_layer.GetNextFeature()
            if first_feature:
                geom_type = first_feature.GetGeometryRef().GetGeometryType()
            else:
                raise Exception(
                    "Could not determine geometry type for WFS layer, no features exist."
                )

        # Initialize output layer
        output_layer = self.output_data_source.CreateLayer(
            layer_name,
            srs=input_layer.GetSpatialRef(),
            geom_type=geom_type,
        )
        if output_layer is None:
            raise Exception(
                f"Could not create layer {layer_name} in output data source."
            )

        # Create all remaining non-geometry columns
        for i in range(input_layer_defn.GetFieldCount()):
            field_defn = input_layer_defn.GetFieldDefn(i)
            output_layer.CreateField(field_defn)

        # Copy features from input to output layer
        for feature in input_layer:
            output_layer.CreateFeature(feature)

        # Cleanup
        self.output_data_source = None
        wfs_data_source = None

        ogr.DontUseExceptions()

    async def fetch_mvt(self) -> None:
        """Fetch data from MVT service and save to disk."""
        # TODO: Implement MVT fetching
        pass


class OGRFileHandling:
    def __init__(
        self, async_session: AsyncSession, user_id: UUID, file_path: str
    ) -> None:
        self.async_session = async_session
        self.user_id = user_id
        self.file_path = file_path
        self.folder_path = os.path.dirname(file_path)
        self.file_ending = os.path.splitext(os.path.basename(file_path))[-1][1:]
        self.file_name = os.path.splitext(os.path.basename(file_path))[0]
        self.method_match_validate = {
            FileUploadType.csv.value: self.validate_csv,
            FileUploadType.xlsx.value: self.validate_xlsx,
            FileUploadType.zip.value: self.validate_shapefile,
            FileUploadType.gpkg.value: self.validate_gpkg,
            FileUploadType.geojson.value: self.validate_geojson,
            FileUploadType.kml.value: self.validate_kml,
        }
        if self.file_ending == FileUploadType.csv.value:
            self.driver_name = OgrDriverType[FileUploadType.xlsx.value].value
        else:
            self.driver_name = OgrDriverType[self.file_ending].value

    def validate_ogr(self, file_path: str) -> Dict[str, Any]:
        """Validate using ogr and get valid attributes."""

        # Get driver
        driver = ogr.GetDriverByName(self.driver_name)

        # Open the file using OGR
        data_source = driver.Open(file_path)
        if data_source is None:
            return {
                "msg": "Could not open the file",
                "status": JobStatusType.failed.value,
            }

        # Count the number of layers
        layer_count = data_source.GetLayerCount()

        # Validate that there is only one layer
        if layer_count != 1:
            return {
                "msg": "File must contain exactly one layer.",
                "status": JobStatusType.failed.value,
            }

        # Get Layer and check types
        layer = data_source.GetLayer(0)

        # Check if layer has a geometry
        if layer.GetLayerDefn().GetGeomFieldCount() == 1:
            # Check if CRS is give other no conversion can be done later
            spatial_ref = layer.GetSpatialRef()

            if spatial_ref is None:
                return {
                    "msg": "Could not determine Coordinate Reference System (CRS).",
                    "status": JobStatusType.failed.value,
                }

        field_type_res = self.check_field_types(layer)

        # Close the datasource
        data_source = None

        return {"file_path": file_path, **field_type_res}

    def get_layer_extent(self, layer: ogr.Layer) -> str:
        """Get layer extent in EPSG:4326."""
        # Get the original extent
        minx, maxx, miny, maxy = layer.GetExtent()

        # Define the source SRS
        source_srs = layer.GetSpatialRef()

        # Define the target SRS (EPSG:4326)
        target_srs = osr.SpatialReference()
        target_srs.ImportFromEPSG(4326)

        # Create a coordinate transformation
        transform = osr.CoordinateTransformation(source_srs, target_srs)

        # Transform the coordinates
        min_point = ogr.Geometry(ogr.wkbPoint)
        min_point.AddPoint(minx, miny)
        min_point.Transform(transform)

        max_point = ogr.Geometry(ogr.wkbPoint)
        max_point.AddPoint(maxx, maxy)
        max_point.Transform(transform)

        # Get the transformed coordinates
        miny_transformed, minx_transformed, _ = min_point.GetPoint()
        maxy_transformed, maxx_transformed, _ = max_point.GetPoint()

        # Create a Multipolygon from the extent
        multipolygon_wkt = f"MULTIPOLYGON((({minx_transformed} {miny_transformed}, {minx_transformed} {maxy_transformed}, {maxx_transformed} {maxy_transformed}, {maxx_transformed} {miny_transformed}, {minx_transformed} {miny_transformed})))"
        return multipolygon_wkt

    def check_field_types(self, layer: ogr.Layer) -> Dict[str, Any]:
        """Check if field types are valid and label if too many columns where specified."""

        field_types: Dict[str, Any] = {
            "valid": {},
            "unvalid": {},
            "overflow": {},
            "geometry": {},
        }
        layer_def = layer.GetLayerDefn()

        if layer_def.GetGeomFieldCount() == 1:
            # Get geometry type of layer to upload to specify target table
            if layer_def.GetGeomType() != ogr.wkbUnknown:
                geometry_type = layer_def.GetGeomType()
            else:
                first_feature = layer.GetNextFeature()
                if first_feature:
                    geometry_type = first_feature.GetGeometryRef().GetGeometryType()
                else:
                    raise Exception(
                        "Could not determine geometry type for layer, no features exist."
                    )
            geometry_type = ogr.GeometryTypeToName(geometry_type).replace(" ", "_")

            # Strip the "Measured " from beginning of the the geometry type name
            geometry_type = geometry_type.replace("Measured_", "")
            # Strip "Z", "M", "ZM" or "25D" from the end of the geometry type name
            geometry_type = re.sub(r"(Z|M|ZM|25D)$", "", geometry_type)
            if geometry_type not in SupportedOgrGeomType.__members__:
                return {
                    "msg": "Geometry type not supported.",
                    "status": JobStatusType.failed.value,
                }

            # Save geometry type and geometry column name
            if layer_def.GetGeomFieldDefn(0).GetName() == "":
                field_types["geometry"]["column_name"] = "wkb_geometry"
            else:
                field_types["geometry"]["column_name"] = layer_def.GetGeomFieldDefn(
                    0
                ).GetName()
            field_types["geometry"]["type"] = geometry_type
            field_types["geometry"]["extent"] = self.get_layer_extent(layer)
            field_types["geometry"]["srs"] = (
                "EPSG:" + layer.GetSpatialRef().GetAuthorityCode(None)
            )

        for i in range(layer_def.GetFieldCount()):
            field_def = layer_def.GetFieldDefn(i)
            field_name = field_def.GetName().lower()
            field_type_code = field_def.GetType()
            field_sub_type_code = field_def.GetSubType()
            field_type = field_def.GetFieldTypeName(field_type_code)
            field_sub_type = ogr.GetFieldSubTypeName(field_sub_type_code)

            # Get field type from OgrPostgresType enum if exists
            if field_type in OgrPostgresType.__members__:
                if field_sub_type in OgrPostgresSubType.__members__:
                    field_type_pg = OgrPostgresSubType[field_sub_type].value
                else:
                    field_type_pg = OgrPostgresType[field_type].value
            else:
                field_type_pg = None

            # Check if field type is defined
            if field_type_pg is None:
                field_types["unvalid"][field_name] = field_type
                continue
            # Create array for field names of respective type if not already existing
            if field_type_pg not in field_types["valid"].keys():
                field_types["valid"][field_type_pg] = []

            # Check if number of specified field excesses the maximum specified number
            if (
                NumberColumnsPerType[field_type_pg].value
                > len(field_types["valid"][field_type_pg])
                and field_name not in field_types["valid"][field_type_pg]
            ):
                field_types["valid"][field_type_pg].append(field_name)

            # Place fields that are exceeding the maximum number of columns or if the column name was already specified.
            elif (
                NumberColumnsPerType[field_type_pg]
                <= len(field_types["valid"][field_type_pg])
                or field_name in field_types["valid"][field_type_pg]
            ):
                field_types["overflow"][field_type_pg] = field_name

        return {"data_types": field_types}

    def validate_csv(self) -> Dict[str, Any]:
        """Validate if CSV."""

        # Read file in binary mode to check if header exists
        with open(self.file_path, "rb") as f:
            sniffer = csv.Sniffer()
            sample_bytes = 1024
            sample = f.read(sample_bytes)

            if not sniffer.has_header(sample.decode("utf-8")):
                return {
                    "msg": "CSV is not well-formed: Missing header.",
                    "status": JobStatusType.failed.value,
                }

        # Read file in text mode
        with open(self.file_path, "r") as f:
            # Get header
            csv_reader = csv.reader(f)
            header = next(csv_reader)

            if any(not col for col in header):
                return {
                    "msg": "CSV is not well-formed: Header contains empty values.",
                    "status": JobStatusType.failed.value,
                }

        # Load in df to get data types
        df = pd.read_csv(self.file_path)
        # Save in XLSX to keep data types
        df.to_excel(self.file_path + ".xlsx", index=False)
        return self.validate_ogr(self.file_path + ".xlsx")

    def validate_xlsx(self) -> Dict[str, Any]:
        """Validate if XLSX is well-formed."""
        # Load workbook
        wb = load_workbook(filename=self.file_path, read_only=True)

        # Check if only one sheet is present
        if len(wb.sheetnames) != 1:
            return {
                "msg": "XLSX is not well-formed: More than one sheet is present.",
                "status": JobStatusType.failed.value,
            }

        sheet = wb.active
        # Check header
        header = [cell.value for cell in sheet[1]]
        if any(value is None for value in header):
            return {
                "msg": "XLSX is not well-formed: Header contains empty values.",
                "status": JobStatusType.failed.value,
            }

        return self.validate_ogr(self.file_path)

    def validate_shapefile(self) -> Dict[str, Any]:
        """Validate if ZIP contains a valid shapefile."""
        with zipfile.ZipFile(self.file_path) as zip_ref:
            # List all file names in the zip file
            file_names = zip_ref.namelist()
            # Remove directories from the list of file names
            file_names = [f for f in file_names if not f.endswith("/")]

            # Check for required shapefile components
            extensions = [".shp", ".shx", ".dbf", ".prj"]
            valid_files = []
            for ext in extensions:
                matching_files = [f for f in file_names if f.endswith(ext)]
                if len(matching_files) != 1:
                    return {
                        "msg": f"ZIP must contain exactly one {ext} file.",
                        "status": JobStatusType.failed.value,
                    }
                valid_files.append(matching_files[0])

            # Check if the main shapefile components share the same base name
            base_name = os.path.splitext(valid_files[0])[0]
            if any(f"{base_name}{ext}" not in valid_files for ext in extensions):
                return {
                    "msg": "All main shapefile components (.shp, .shx, .dbf, .prj) must share the same base name.",
                    "status": JobStatusType.failed.value,
                }

        # Unzip file in temporary directory
        zip_dir = os.path.join(
            os.path.dirname(self.file_path),
            os.path.basename(self.file_path).split(".")[0],
        )

        # Extra zip file to zip_dir
        with zipfile.ZipFile(self.file_path) as zip_ref:
            zip_ref.extractall(zip_dir)

        return self.validate_ogr(os.path.join(zip_dir, base_name + ".shp"))

    def validate_gpkg(self) -> Dict[str, Any]:
        """Validate geopackage."""
        return self.validate_ogr(self.file_path)

    def validate_geojson(self) -> Dict[str, Any]:
        """Validate geojson."""
        return self.validate_ogr(self.file_path)

    def validate_kml(self) -> Dict[str, Any]:
        """Validate kml."""
        return self.validate_ogr(self.file_path)

    async def validate(self) -> Dict[str, Any]:
        """Validate file before uploading."""

        # Run validation
        result = self.method_match_validate[self.file_ending]()

        if result.get("status") == JobStatusType.failed.value:
            return result

        # Build object for job step status
        msg_text = ""
        if (
            result["data_types"]["unvalid"] == {}
            and result["data_types"]["overflow"] == {}
        ):
            msg_type = MsgType.info
            msg_text = "File is valid."
        else:
            msg_type = MsgType.warning
            if result["data_types"]["unvalid"]:
                msg_text = f"The following attributes are not saved as they could not be mapped to a valid data type: {', '.join(result['data_types']['unvalid'].values())}"
            if result["data_types"]["overflow"]:
                msg_text = (
                    msg_text
                    + f"The following attributes are not saved as they exceed the maximum number of columns per data type: {', '.join(result['data_types']['overflow'].values())}"
                )

        result["msg"] = Msg(type=msg_type, text=msg_text)
        return result

    async def validate_fail(self, folder_path: str) -> None:
        """Delete folder if validation fails."""
        await async_delete_dir(folder_path)

    @job_log(job_step_name="upload")
    async def upload_ogr2ogr(
        self, temp_table_name: str, job_id: UUID
    ) -> Dict[str, Any]:
        """Upload file to database."""

        # Initialize OGR
        ogr.RegisterAll()

        # Setup the input GeoJSON data source
        driver = ogr.GetDriverByName(self.driver_name)
        data_source = driver.Open(self.file_path, 0)
        layer = data_source.GetLayer(0)
        layer_def = layer.GetLayerDefn()

        # Get geometry type of layer to force multipolygon if any polygon type
        geometry_type = ogr.GeometryTypeToName(layer_def.GetGeomType()).replace(
            " ", "_"
        )
        if "polygon" in geometry_type.lower():
            geometry_type = "-nlt MULTIPOLYGON"
        else:
            geometry_type = ""

        # Prepare the ogr2ogr command
        if self.file_ending == FileUploadType.gpkg.value:
            layer_name = layer.GetName()
        else:
            layer_name = None

        # Build CMD command
        cmd = (
            f'ogr2ogr -f "PostgreSQL" "PG:host={settings.POSTGRES_SERVER} dbname={settings.POSTGRES_DB} '
            f'user={settings.POSTGRES_USER} password={settings.POSTGRES_PASSWORD} port={settings.POSTGRES_PORT}" '
            f'"{self.file_path}" '
        )
        if layer_name:
            cmd += f'"{layer_name}" '
        cmd += f'-nln {temp_table_name} -t_srs "EPSG:4326" -progress -dim XY {geometry_type} -unsetFieldWidth'
        try:
            # Run as async task
            task = asyncio.create_task(async_run_command(cmd))
            await task
        except Exception as e:
            raise Ogr2OgrError(sanitize_error_message(str(e)))

        # Close data source
        data_source = None

        # Build object for job step status
        msg = Msg(type=MsgType.info, text="File uploaded.")

        # Delete folder with file
        await async_delete_dir(self.folder_path)
        return {
            "msg": msg,
            "status": JobStatusType.finished.value,
        }

    async def upload_ogr2ogr_fail(self, temp_table_name: str) -> None:
        """Delete folder if ogr2ogr upload fails."""
        await self.validate_fail(self.folder_path)
        await self.async_session.execute(
            text(f"DROP TABLE IF EXISTS {temp_table_name}")
        )
        await self.async_session.commit()

    async def export_ogr2ogr(
        self,
        layer: Layer,
        file_type: TableLayerExportType | FeatureLayerExportType,
        file_name: str,
        sql_query: str,
        crs: str,
    ) -> str:
        """Export file from database."""

        # Initialize OGR
        ogr.RegisterAll()

        # Prepare the ogr2ogr command
        if self.file_ending == FileUploadType.gpkg.value:
            pass
        else:
            pass

        # Create folder if not exists
        await async_delete_dir(
            os.path.join(settings.DATA_DIR, str(self.user_id), str(layer.id))
        )
        if not os.path.exists(os.path.join(settings.DATA_DIR, str(self.user_id))):
            await aos.mkdir(os.path.join(settings.DATA_DIR, str(self.user_id)))
        await aos.mkdir(
            os.path.join(settings.DATA_DIR, str(self.user_id), str(layer.id))
        )
        await aos.mkdir(self.folder_path)

        if layer.type == LayerType.feature.value:
            # Define target CRS
            target_crs = CRS(crs)

            # Get the layer's extent
            minx, miny, maxx, maxy = wkb.loads(layer.extent.desc, hex=True).bounds

            # Check if the layer's extent falls within the bounds of the target CRS
            if not (
                target_crs.area_of_use.west
                <= minx
                <= maxx
                <= target_crs.area_of_use.east
                and target_crs.area_of_use.south
                <= miny
                <= maxy
                <= target_crs.area_of_use.north
            ):
                raise DataOutCRSBoundsError(
                    "The data is outside the bounds of the provided CRS."
                )
            to_crs_flag = f"""-t_srs "{crs}" """
        else:
            to_crs_flag = ""

        # Build CMD command
        sql_query = sql_query.replace('"', '\\"')
        cmd = f"""ogr2ogr -f "{OgrDriverType[file_type.value].value}" "{self.file_path}" PG:"host={settings.POSTGRES_SERVER} dbname={settings.POSTGRES_DB} user={settings.POSTGRES_USER} password={settings.POSTGRES_PASSWORD} port={settings.POSTGRES_PORT}" -sql "{sql_query}" -nln "{layer.name}" {to_crs_flag} -progress"""
        try:
            # Run as async task
            task = asyncio.create_task(async_run_command(cmd))
            await task
        except Exception as e:
            raise Ogr2OgrError(sanitize_error_message(str(e)))

        return self.file_path

    # @timeout(120)
    @job_log(job_step_name="migration")
    async def migrate_target_table(
        self,
        validation_result: Dict[str, Any],
        attribute_mapping: Dict[str, Any],
        temp_table_name: str,
        layer_id: UUID,
        job_id: UUID,
    ) -> Dict[str, Any]:
        """Migrate data from temporary table to target table."""

        data_types = validation_result["data_types"]
        geom_column = data_types["geometry"].get("column_name")

        # Check if table has a geometry if not it is just a normal table
        if geom_column is None:
            target_table = f"{settings.USER_DATA_SCHEMA}.no_geometry_{str(self.user_id).replace('-', '')}"
            select_geom = ""
            insert_geom = ""
            filter_null_geom = ""
        else:
            geometry_type = data_types["geometry"]["type"]
            target_table = f"{settings.USER_DATA_SCHEMA}.{SupportedOgrGeomType[geometry_type].value}_{str(self.user_id).replace('-', '')}"
            select_geom = f"{geom_column} as geom, "
            insert_geom = "geom, "
            filter_null_geom = f"WHERE ST_IsEmpty({geom_column}) IS FALSE"
        select_statement = ""
        insert_statement = ""

        # Build select and insert statement
        for i in attribute_mapping:
            data_type = i.split("_")[0]
            select_statement += f""""{attribute_mapping[i]}"::{data_type} as {i}, """
            insert_statement += f"{i}, "
        select_statement = f"""SELECT {select_statement} {select_geom} '{str(layer_id)}' FROM {temp_table_name} {filter_null_geom}"""

        # Insert data in target table
        await self.async_session.execute(
            text(
                f"INSERT INTO {target_table}({insert_statement} {insert_geom} layer_id) {select_statement}"
            )
        )
        await self.async_session.commit()

        # Delete temporary table
        await self.async_session.execute(
            text(f"DROP TABLE IF EXISTS {temp_table_name}")
        )
        await self.async_session.commit()

        return {
            "msg": Msg(type=MsgType.info, text="Data migrated."),
            "status": JobStatusType.finished.value,
        }

    async def migrate_target_table_fail(
        self,
        validation_result: Dict[str, Any],
        attribute_mapping: Dict[str, Any],
        temp_table_name: str,
        layer_id: UUID,
    ) -> None:
        """Delete folder if ogr2ogr upload fails."""

        # Delete data from user table if already inserted
        data_types = validation_result["data_types"]
        geom_column = data_types["geometry"].get("column_name")

        # Check if table has a geometry if not it is just a normal table
        if geom_column is None:
            target_table = f"{settings.USER_DATA_SCHEMA}.no_geometry_{str(self.user_id).replace('-', '')}"
        else:
            geometry_type = data_types["geometry"]["type"]
            target_table = f"{settings.USER_DATA_SCHEMA}.{SupportedOgrGeomType[geometry_type].value}_{str(self.user_id).replace('-', '')}"

        await self.upload_ogr2ogr_fail(temp_table_name)
        await self.async_session.execute(
            text(f"DELETE FROM {target_table} WHERE layer_id = '{str(layer_id)}'")
        )
        await self.async_session.commit()


async def delete_layer_data(async_session: AsyncSession, layer: Layer) -> None:
    """Delete layer data which is in the user data tables."""

    # Delete layer data
    await async_session.execute(
        text(f"DELETE FROM {layer.table_name} WHERE layer_id = '{layer.id}'")
    )
    await async_session.commit()
